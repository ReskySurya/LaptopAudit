"""
Report: XLSX — Generate spreadsheet Excel untuk database audit.
Menggunakan openpyxl.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import KOMPONEN_FISIK


def generate_xlsx(data: dict, output_path: str):
    """
    Generate Excel spreadsheet dari data audit.
    Satu baris = satu record audit (database-ready format).

    Args:
        data: dict gabungan data manual + scan
        output_path: path file XLSX output
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Data"

    # ── Define columns (header) ───────────────────────
    headers = [
        "Tanggal Audit",
        "PIC Asset",
        "Kode Asset",
        "Jenis Asset",
        "Lama Dibawa",
        "Hostname",
        "Username",
        "OS",
        "OS Build",
        "Vendor",
        "Model",
        "Serial Number",
        "CPU",
        "CPU Cores",
        "CPU Threads",
        "RAM (GB)",
        "IP Address",
        "MAC Address",
        "Battery (%)",
        "Charging Status",
        "Design Capacity",
        "Full Charge Capacity",
        "Battery Health",
        "Firewall Status",
        "Antivirus",
        "Antivirus Status",
    ]

    # Tambah kolom untuk setiap komponen fisik
    for komponen in KOMPONEN_FISIK:
        headers.append(f"Kondisi {komponen}")

    # Tambah kolom storage
    headers.append("Storage Info")
    headers.append("Jumlah Software Terinstall")

    # ── Style header ──────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write header
    for col, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ── Build data row ────────────────────────────────
    bat = data.get("battery", {})
    net = data.get("network", {})
    sec = data.get("security", {})
    matriks = data.get("matriks_fisik", {})

    # Storage info sebagai string
    storage = data.get("storage", [])
    storage_str = "; ".join(
        f"{s.get('drive', '?')}: {s.get('total_gb', '?')}GB total, "
        f"{s.get('free_gb', '?')}GB free"
        for s in storage
    ) if storage else "N/A"

    row_data = [
        data.get("tanggal", ""),
        data.get("pic_asset", ""),
        data.get("kode_asset", ""),
        data.get("jenis_asset", ""),
        data.get("lama_dibawa", ""),
        data.get("hostname", ""),
        data.get("username", ""),
        data.get("os_name", ""),
        data.get("os_build", ""),
        data.get("vendor", ""),
        data.get("model", ""),
        data.get("serial_number", ""),
        data.get("cpu_name", ""),
        data.get("cpu_cores", ""),
        data.get("cpu_threads", ""),
        data.get("ram_total_gb", ""),
        ", ".join(net.get("ip_addresses", ["-"])),
        ", ".join(net.get("mac_addresses", ["-"])),
        bat.get("percent", "N/A"),
        bat.get("charging_status", "N/A"),
        bat.get("design_capacity_wh", "N/A"),
        bat.get("full_charge_capacity_wh", "N/A"),
        bat.get("health_percent", "N/A"),
        sec.get("firewall_status", "N/A"),
        sec.get("antivirus_name", "N/A"),
        sec.get("antivirus_status", "N/A"),
    ]

    # Kondisi fisik
    for komponen in KOMPONEN_FISIK:
        row_data.append(matriks.get(komponen, "-"))

    row_data.append(storage_str)
    row_data.append(len(data.get("software_list", [])))

    # Write data row
    data_font = Font(name="Calibri", size=10)
    data_alignment = Alignment(vertical="center", wrap_text=True)

    for col, value in enumerate(row_data, start=1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

        # Warnai kondisi fisik
        if isinstance(value, str):
            if value == "Baik":
                cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                cell.font = Font(name="Calibri", size=10, color="FFFFFF", bold=True)
            elif value == "Rusak":
                cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                cell.font = Font(name="Calibri", size=10, color="FFFFFF", bold=True)
            elif value == "Lecet":
                cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
                cell.font = Font(name="Calibri", size=10, color="FFFFFF", bold=True)

    # ── Auto-width columns ────────────────────────────
    for col in range(1, len(headers) + 1):
        max_length = 0
        for row in range(1, 3):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max(max_length + 3, 12), 45)
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Sheet 2: Software List ────────────────────────
    sw_list = data.get("software_list", [])
    if sw_list:
        ws2 = wb.create_sheet("Software List")

        # Header
        ws2.cell(row=1, column=1, value="No").font = header_font
        ws2.cell(row=1, column=1).fill = header_fill
        ws2.cell(row=1, column=1).border = thin_border
        ws2.cell(row=1, column=2, value="Nama Software").font = header_font
        ws2.cell(row=1, column=2).fill = header_fill
        ws2.cell(row=1, column=2).border = thin_border

        for i, sw_name in enumerate(sw_list, start=1):
            ws2.cell(row=i + 1, column=1, value=i).border = thin_border
            ws2.cell(row=i + 1, column=2, value=sw_name).border = thin_border

        ws2.column_dimensions["A"].width = 8
        ws2.column_dimensions["B"].width = 50
        ws2.freeze_panes = "A2"

    wb.save(output_path)
